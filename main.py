from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command, Text
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InputFile
from datetime import datetime
from io import BytesIO
from win32com import client


import os
import matplotlib.pyplot as plt
import openpyxl
from fpdf import FPDF

from Markups import menu_markup, back_cancel_markup
from db import BotDB

Database = BotDB(db_file="Tapo.db")

# Замените 'YOUR_BOT_TOKEN' на токен вашего бота Telegram
TOKEN = '5791370863:AAEsAXyZ1-8z3lNeC8FnP7hR7JV5a578HTc'

# Инициализация бота и диспетчера
bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)


# Обработчик команды /start
@dp.message_handler(Command("start"))
async def cmd_start(message: types.Message):
    Database.add_user(message.from_user.id, message.from_user.username)
    await bot.send_message(message.chat.id,"Приветствую тебя, дорогой пользователь. С помощью этого бота ты сможешь определить код своей энерги и полуть подробный расклад на год\nСтоимость прогноза - 20 $",reply_markup=menu_markup)

class Prognoz(StatesGroup):
    Date = State()
    Name = State()

@dp.message_handler(Text(equals="Заказать прогноз"))
async def process_birthday(message: types.Message):
    await bot.send_message(message.chat.id, 'Введите дату рождения (в формате ДД.ММ.ГГГГ):',reply_markup=back_cancel_markup)
    await Prognoz.Date.set()



@dp.message_handler(state=Prognoz.Date)
async def process_birthday(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await bot.send_message(message.chat.id, "Меню:", reply_markup=menu_markup)
        await state.finish()
    if message.text == "Отмена":
        await bot.send_message(message.chat.id, "Меню:", reply_markup=menu_markup)
        await state.finish()
    else:
        birthday = message.text
        balance = Database.get_balance(message.from_user.id)
        if int(balance) >= 20:
            try:
                # Загрузка файла Excel
                wb = openpyxl.load_workbook('1.xlsx')
                sheet = wb['Данные']

                # Запрос даты рождения у пользователя

                day, month, year = birthday.split('.')

                # Вставка даты рождения в ячейку С6
                sheet['C5'] = birthday
                # Закрытие файла Excel
                wb.save('1.xlsx')
                wb.close()
                await bot.send_message(message.from_user.id, "Введите ваше имя и фамилию")
                await Prognoz.Name.set()
            except:
                # Обрабатываем ошибку, если дата введена неверно
                await message.reply("Неверный формат даты. Введите дату в формате ДД.ММ.ГГГГ")
        else:
            await bot.send_message(message.from_user.id,"Недостаточно средств!")

@dp.message_handler(state=Prognoz.Name)
async def process_birthday(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await bot.send_message(message.chat.id, 'Введите дату рождения (в формате ДД.ММ.ГГГГ):',reply_markup=back_cancel_markup)
        await Prognoz.Date.set()
    if message.text == "Отмена":
        await bot.send_message(message.chat.id, "Меню:", reply_markup=menu_markup)
        await state.finish()
    else:
        name = message.text
        # Загрузка файла Excel
        wb = openpyxl.load_workbook('1.xlsx')
        sheet = wb['Данные']

        sheet['C3'] = name

        sheets = wb.sheetnames

        wb.save('1.xlsx')
        wb.close()

        # # Определяем последние два листа
        # last_two_sheets = ["Прогноз год","Прогноз месяц","Расшифровка по дням","Расшифровка по месяцу"]
        #
        # # Создаем новый PDF-документ
        # pdf = FPDF()
        #
        #
        #
        # # Проходимся по последним двум листам
        # for sheet_name in last_two_sheets:
        #     sheet = wb[sheet_name]
        #
        #     # Добавляем новую страницу в PDF-документ
        #     pdf.add_page()
        #
        #     # Устанавливаем шрифт и размер текста
        #     pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
        #     pdf.set_font('DejaVu', '', 12)
        #
        #     # Получаем данные из листа и добавляем их в PDF-документ
        #     for row in sheet.iter_rows(values_only=True):
        #         pdf.cell(200, 10, txt=str(row), ln=1)

        xlApp = client.Dispatch("Excel.Application")
        books = xlApp.Workbooks.Open('D:\Projects\Таро\Tapo new_version\\1.xlsx')
        ws = books.Worksheets[0]
        ws.Visible = 1
        ws.ExportAsFixedFormat(0, 'D:\Projects\Таро\Tapo new_version\\output.pdf')

        # Сохраняем PDF-документ
        pdf_file_path = 'output.pdf'

        # Отправляем PDF-файл пользователю
        with open(pdf_file_path, 'rb') as file:
            await bot.send_document(message.chat.id, document=InputFile(file, filename='output.pdf'))

        # Удаляем PDF-файл после отправки
        os.remove(pdf_file_path)

        # except:
        #     # Обрабатываем ошибку, если дата введена неверно
        #     await message.reply("Введите ваше Имя и Фамилию")





# Обработчик неизвестных команд и текстовых сообщений
@dp.message_handler()
async def unknown_message(message: types.Message):
    await message.reply("Неизвестная команда. Введите /start для начала.")


if __name__ == "__main__":
    # Запуск бота
    from aiogram import executor

    # Регистрация хэндлеров
    dp.register_message_handler(cmd_start, commands="start")
    dp.register_message_handler(unknown_message)

    executor.start_polling(dp, skip_updates=True)
